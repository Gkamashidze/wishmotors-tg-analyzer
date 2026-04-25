import html
import logging
from datetime import date, datetime, timezone
from io import BytesIO
from typing import Optional

import openpyxl
import pytz
from aiogram import F, Router, Bot
from aiogram.enums import ParseMode
from aiogram.types import InlineKeyboardButton, InlineKeyboardMarkup, Message

import config
from bot.handlers import InTopic, IsAdmin
from bot.handlers.topic_messages import topic_nisia_kb, topic_sale_kb
from bot.parsers.message_parser import (
    ParsedSale,
    parse_batch_sales,
    parse_dual_sale_message,
    parse_sale_message,
    sanitize_oem,
)
from bot.reports.formatter import (
    format_batch_confirmation,
    format_sale_confirmation,
    format_return_confirmation,
    format_topic_sale,
)
from database.db import Database
from database.models import ProductRow

logger = logging.getLogger(__name__)
sales_router = Router(name="sales")

_PARSE = ParseMode.HTML
_MAX_IMPORT_ROWS = 2_000  # Safety limit per Excel import

_DATE_FORMATS = ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d")


def _parse_backdate(raw: object) -> Optional[datetime]:
    """Return a UTC-aware datetime from an Excel cell value, or None if absent/unparseable."""
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.replace(tzinfo=timezone.utc) if raw.tzinfo is None else raw.astimezone(timezone.utc)
    if isinstance(raw, date):
        return datetime(raw.year, raw.month, raw.day, tzinfo=timezone.utc)
    s = str(raw).strip()
    if not s:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


def _delete_keyboard(sale_id: int) -> InlineKeyboardMarkup:
    """Single delete button for one sale confirmation."""
    return InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text=f"🗑 წაშლა #{sale_id}", callback_data=f"ds:{sale_id}")
    ]])


def _delete_keyboard_batch(sale_ids: list[int]) -> InlineKeyboardMarkup:
    """Delete buttons for a batch confirmation — pairs per row."""
    rows = []
    for i in range(0, len(sale_ids), 2):
        row = [
            InlineKeyboardButton(text=f"🗑 #{sid}", callback_data=f"ds:{sid}")
            for sid in sale_ids[i:i + 2]
        ]
        rows.append(row)
    return InlineKeyboardMarkup(inline_keyboard=rows)


# ─── Sales topic: text messages ───────────────────────────────────────────────

@sales_router.message(InTopic(config.SALES_TOPIC_ID), IsAdmin(), F.text)
async def handle_sales_text(message: Message, db: Database) -> None:
    text = (message.text or "").strip()
    try:
        # Multi-line → batch handler (one customer, many items)
        if "\n" in text:
            await _handle_batch_sales(message, db, text)
            return

        # Dual-product format: "008b03 და 108000 1-1ც ჯამში 100₾" → two separate sales
        dual = parse_dual_sale_message(text)
        if dual is not None:
            await _handle_dual_sale(message, db, dual)
            return

        parsed = parse_sale_message(text)

        if not parsed:
            await db.log_parse_failure(config.SALES_TOPIC_ID, text)
            return

        raw = parsed.raw_product

        product = await db.get_product_by_oem(raw)
        if not product and raw:
            product = await db.get_product_by_partial_oem(raw)
        if not product:
            product = await db.get_product_by_name(raw)

        if not product:
            if parsed.is_return:
                await message.bot.send_message(
                    chat_id=message.from_user.id,
                    text=f"⚠️ პროდუქტი <b>{html.escape(raw)}</b> ვერ მოიძებნა. დაბრუნება ვერ ჩაიწერება.",
                    parse_mode=_PARSE,
                )
                return
            await _record_sale_freeform(message, db, raw, parsed)
            return

        if parsed.is_return:
            await _record_return(message, db, product, parsed)
        else:
            await _record_sale(message, db, product, parsed)
    except Exception:
        logger.exception("Unexpected error in handle_sales_text: %r", text)
        await message.bot.send_message(
            chat_id=message.from_user.id,
            text="❌ სისტემური შეცდომა. გთხოვთ, სცადოთ ხელახლა.",
            parse_mode=_PARSE,
        )


async def _handle_batch_sales(message: Message, db: Database, text: str) -> None:
    """Process a multi-line message — each line is a separate sale, customer shared."""
    customer_name, parsed_list = parse_batch_sales(text)

    if not parsed_list:
        await db.log_parse_failure(config.SALES_TOPIC_ID, text)
        return

    results = []
    failed_lines = []
    grand_total = 0.0
    partial_payment: float = 0.0  # cash already received ("მომცა X დარჩა Y")

    raw_lines = [ln.strip() for ln in text.split("\n") if ln.strip()]
    offset = 1 if customer_name else 0

    for i, item_group in enumerate(parsed_list):
        if item_group is None:
            line_idx = offset + i
            failed_lines.append(raw_lines[line_idx] if line_idx < len(raw_lines) else "?")
            continue

        for parsed in item_group:
            # Split-payment line ("მომცა 300 დარჩა 100") — apply after all sales recorded
            if parsed.is_split_payment:
                partial_payment += parsed.split_paid
                continue

            raw = parsed.raw_product
            product = await db.get_product_by_oem(raw) if raw else None
            if not product and raw:
                product = await db.get_product_by_partial_oem(raw)
            if not product and raw:
                product = await db.get_product_by_name(raw)

            sale_id, _ = await db.create_sale(
                product_id=product["id"] if product else None,
                quantity=parsed.quantity,
                unit_price=parsed.price,
                payment_method=parsed.payment_method,
                seller_type=parsed.seller_type,
                customer_name=parsed.customer_name or None,
                notes=raw if not product else None,
            )
            grand_total += parsed.quantity * parsed.price
            results.append((parsed, product, sale_id))

    if not results:
        await db.log_parse_failure(config.SALES_TOPIC_ID, text)
        return

    # Apply cash partial payment against the customer's new credit sales
    if partial_payment > 0 and customer_name:
        await db.apply_partial_payment(customer_name, partial_payment)

    sale_ids = [sale_id for _, _, sale_id in results]
    await message.bot.send_message(
        chat_id=message.from_user.id,
        text=format_batch_confirmation(customer_name, results, grand_total, failed_lines),
        parse_mode=_PARSE,
        reply_markup=_delete_keyboard_batch(sale_ids),
    )


async def _handle_dual_sale(message: Message, db: Database, dual: list) -> None:
    """Record two sales from a single-line dual entry (e.g. '008b03 და 108000 1-1ც ჯამში 100₾')."""
    results = []
    grand_total = 0.0
    customer_name = dual[0].customer_name or None

    for parsed in dual:
        raw = parsed.raw_product
        product = await db.get_product_by_oem(raw) if raw else None
        if not product and raw:
            product = await db.get_product_by_partial_oem(raw)
        if not product and raw:
            product = await db.get_product_by_name(raw)

        sale_id, _ = await db.create_sale(
            product_id=product["id"] if product else None,
            quantity=parsed.quantity,
            unit_price=parsed.price,
            payment_method=parsed.payment_method,
            seller_type=parsed.seller_type,
            customer_name=parsed.customer_name or None,
            notes=raw if not product else None,
        )
        grand_total += parsed.quantity * parsed.price
        results.append((parsed, product, sale_id))

    sale_ids = [sale_id for _, _, sale_id in results]
    await message.bot.send_message(
        chat_id=message.from_user.id,
        text=format_batch_confirmation(customer_name, results, grand_total, []),
        parse_mode=_PARSE,
        reply_markup=_delete_keyboard_batch(sale_ids),
    )


async def _record_sale(message: Message, db: Database, product: ProductRow, parsed: ParsedSale) -> None:
    sale_id, new_stock = await db.create_sale(
        product_id=product["id"],
        quantity=parsed.quantity,
        unit_price=parsed.price,
        payment_method=parsed.payment_method,
        seller_type=parsed.seller_type,
        customer_name=parsed.customer_name or None,
    )
    low = new_stock <= product["min_stock"]

    await message.bot.send_message(
        chat_id=message.from_user.id,
        text=format_sale_confirmation(
            product_name=product["name"],
            qty=parsed.quantity,
            price=parsed.price,
            payment=parsed.payment_method,
            seller_type=parsed.seller_type,
            customer_name=parsed.customer_name,
            new_stock=new_stock,
            low_stock=low,
            sale_id=sale_id,
        ),
        parse_mode=_PARSE,
        reply_markup=_delete_keyboard(sale_id),
    )

    # Mirror to topic and save message_id for later deletion
    topic_id = config.NISIAS_TOPIC_ID if parsed.payment_method == "credit" else config.SALES_TOPIC_ID
    kb = topic_nisia_kb(sale_id) if parsed.payment_method == "credit" else topic_sale_kb(sale_id)
    try:
        topic_msg = await message.bot.send_message(
            chat_id=config.GROUP_ID,
            message_thread_id=topic_id,
            text=format_topic_sale(
                product_name=product["name"],
                qty=parsed.quantity,
                price=parsed.price,
                payment=parsed.payment_method,
                sale_id=sale_id,
                customer_name=parsed.customer_name or None,
                oem_code=product.get("oem_code"),
            ),
            parse_mode=_PARSE,
            reply_markup=kb,
        )
        await db.update_sale_topic_message(sale_id, topic_id, topic_msg.message_id)
    except Exception as _te:
        logger.warning("Failed to post sale to topic: %s", _te)

    if low:
        logger.warning(
            "Low stock alert: %s — %d units remaining", product["name"], new_stock
        )

    if new_stock == 0:
        await db.create_order(
            product_id=product["id"],
            quantity_needed=product["min_stock"],
            notes=f"ავტო: {product['name']} — 0-ზე ჩამოვიდა",
        )
        await message.bot.send_message(
            chat_id=message.from_user.id,
            text=(
                f"🔔 <b>ავტოშეკვეთა!</b>\n"
                f"📦 {product['name']} — საწყობი 0-ზე ჩამოვიდა.\n"
                f"📋 შეკვეთა ავტომატურად შეიქმნა ({product['min_stock']}ც)."
            ),
            parse_mode=_PARSE,
        )


async def _record_sale_freeform(
    message: Message, db: Database, product_name: str, parsed: ParsedSale
) -> None:
    """პროდუქტი ბაზაში არ არსებობს — გაყიდვა ჩაიწერება notes-ით, stock ცვლილების გარეშე."""
    sale_id, _ = await db.create_sale(
        product_id=None,
        quantity=parsed.quantity,
        unit_price=parsed.price,
        payment_method=parsed.payment_method,
        seller_type=parsed.seller_type,
        customer_name=parsed.customer_name or None,
        notes=product_name,
    )

    await message.bot.send_message(
        chat_id=message.from_user.id,
        text=format_sale_confirmation(
            product_name=product_name,
            qty=parsed.quantity,
            price=parsed.price,
            payment=parsed.payment_method,
            seller_type=parsed.seller_type,
            customer_name=parsed.customer_name,
            new_stock=None,
            low_stock=False,
            sale_id=sale_id,
            unknown_product=True,
        ),
        parse_mode=_PARSE,
        reply_markup=_delete_keyboard(sale_id),
    )

    topic_id = config.NISIAS_TOPIC_ID if parsed.payment_method == "credit" else config.SALES_TOPIC_ID
    kb = topic_nisia_kb(sale_id) if parsed.payment_method == "credit" else topic_sale_kb(sale_id)
    try:
        topic_msg = await message.bot.send_message(
            chat_id=config.GROUP_ID,
            message_thread_id=topic_id,
            text=format_topic_sale(
                product_name=product_name,
                qty=parsed.quantity,
                price=parsed.price,
                payment=parsed.payment_method,
                sale_id=sale_id,
                customer_name=parsed.customer_name or None,
                unknown_product=True,
            ),
            parse_mode=_PARSE,
            reply_markup=kb,
        )
        await db.update_sale_topic_message(sale_id, topic_id, topic_msg.message_id)
    except Exception as _te:
        logger.warning("Failed to post freeform sale to topic: %s", _te)


async def _record_return(message: Message, db: Database, product: ProductRow, parsed: ParsedSale) -> None:
    refund = parsed.price * parsed.quantity

    _return_id, new_stock = await db.create_return(
        product_id=product["id"],
        quantity=parsed.quantity,
        refund_amount=refund,
        notes="დაბრუნება",
    )

    await message.bot.send_message(
        chat_id=message.from_user.id,
        text=format_return_confirmation(
            product_name=product["name"],
            qty=parsed.quantity,
            refund=refund,
            new_stock=new_stock,
        ),
        parse_mode=_PARSE,
    )


# ─── Sales topic: Excel historical import ────────────────────────────────────

def _parse_import_date(val: object, tz: pytz.BaseTzInfo) -> datetime:
    """Accept Excel date objects or strings (DD.MM.YYYY / YYYY-MM-DD / DD/MM/YYYY)."""
    if isinstance(val, datetime):
        return tz.localize(val.replace(tzinfo=None))
    if isinstance(val, date):
        return tz.localize(datetime(val.year, val.month, val.day))
    s = str(val).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return tz.localize(datetime.strptime(s, fmt))
        except ValueError:
            continue
    raise ValueError(f"თარიღი ვერ წაიკითხა: {s}")


def _parse_import_payment(val: object) -> str:
    v = str(val).lower().strip()
    if any(k in v for k in ("გადარიცხვა", "დარიცხა", "transfer", "ბარათი", "კარტი", "card")):
        return "transfer"
    if any(k in v for k in ("ნისია", "credit")):
        return "credit"
    return "cash"


@sales_router.message(InTopic(config.SALES_TOPIC_ID), IsAdmin(), F.document)
async def handle_sales_import_excel(message: Message, bot: Bot, db: Database) -> None:
    doc = message.document
    if not doc or not doc.file_name:
        return
    if not doc.file_name.lower().endswith((".xlsx", ".xls")):
        await message.bot.send_message(
            chat_id=message.from_user.id,
            text="❌ გთხოვთ Excel ფაილი (.xlsx) გამოაგზავნოთ.",
            parse_mode=_PARSE,
        )
        return

    if doc.file_size and doc.file_size > config.MAX_EXCEL_BYTES:
        mb = config.MAX_EXCEL_BYTES // (1024 * 1024)
        await message.bot.send_message(
            chat_id=message.from_user.id,
            text=f"❌ ფაილი ძალიან დიდია. მაქსიმალური ზომა: <b>{mb} MB</b>.",
            parse_mode=_PARSE,
        )
        return

    await message.bot.send_message(
        chat_id=message.from_user.id,
        text="⏳ გაყიდვების იმპორტი მუშავდება...",
        parse_mode=_PARSE,
    )

    file_info = await bot.get_file(doc.file_id)
    buf = BytesIO()
    await bot.download_file(file_info.file_path, destination=buf)
    buf.seek(0)

    try:
        wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
        ws = wb.active
    except Exception as exc:
        logger.error("Sales import Excel parse error: %s", exc)
        await message.bot.send_message(
            chat_id=message.from_user.id,
            text=(
                "❌ ფაილი ვერ წაიკითხა.\n"
                "სვეტები: <b>თარიღი | პროდუქტი/OEM | რაოდენობა | ფასი | გადახდა</b>"
            ),
            parse_mode=_PARSE,
        )
        return

    tz = pytz.timezone(config.TIMEZONE)
    imported = 0
    errors: list[str] = []
    data_rows = 0

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or row[0] is None:
            continue
        data_rows += 1
        if data_rows > _MAX_IMPORT_ROWS:
            errors.append(f"⚠️ ლიმიტი: {_MAX_IMPORT_ROWS} სტრიქონი. დანარჩენი გამოტოვდა.")
            break
        try:
            sold_at = _parse_import_date(row[0], tz)
            raw_product = str(row[1]).strip() if row[1] is not None else ""
            quantity = int(row[2])
            unit_price = float(row[3])
            payment = _parse_import_payment(row[4]) if len(row) > 4 and row[4] else "credit"

            if not raw_product or quantity <= 0 or unit_price < 0:
                raise ValueError("ცარიელი ან არასწორი მნიშვნელობა")

            product = await db.get_product_by_oem(raw_product)
            if not product:
                product = await db.get_product_by_partial_oem(raw_product)
            if not product:
                product = await db.get_product_by_name(raw_product)

            await db.import_sale(
                product_id=product["id"] if product else None,
                quantity=quantity,
                unit_price=unit_price,
                payment_method=payment,
                sold_at=sold_at,
                notes=raw_product if not product else None,
            )
            imported += 1
        except Exception as exc:
            errors.append(f"სტრიქონი {i}: {exc}")
            logger.warning("Sales import row %d error: %s", i, exc)

    summary = f"✅ <b>იმპორტი დასრულდა!</b>\n📊 ჩაიწერა: <b>{imported}</b> გაყიდვა"
    if errors:
        preview = "\n".join(errors[:5])
        if len(errors) > 5:
            preview += f"\n... და კიდევ {len(errors) - 5}"
        summary += f"\n⚠️ გამოტოვებული ({len(errors)}):\n<code>{html.escape(preview)}</code>"

    await message.bot.send_message(
        chat_id=message.from_user.id,
        text=summary,
        parse_mode=_PARSE,
    )


# ─── Inventory topic: Excel batch receipts (WAC + ledger posting) ─────────────
# Expected columns (header row skipped): თარიღი | OEM | სახელი | მარაგი | ერთეული | ფასი
# Each data row is posted as an inventory receipt: products.current_stock is
# incremented, one inventory_batches row is created, and the ledger gets a
# balanced pair of entries (DR Inventory / CR Accounts payable) so WAC can be
# derived on demand from the inventory_batches table.

@sales_router.message(InTopic(config.STOCK_TOPIC_ID), IsAdmin(), F.document)
async def handle_inventory_upload(message: Message, bot: Bot, db: Database) -> None:
    doc = message.document
    if not doc or not doc.file_name:
        return

    if not doc.file_name.lower().endswith((".xlsx", ".xls")):
        await message.bot.send_message(
            chat_id=message.from_user.id,
            text=(
                "❌ გთხოვთ Excel ფაილი (.xlsx) გამოაგზავნოთ.\n"
                "სვეტები: <b>თარიღი | OEM | სახელი | მარაგი | ერთეული | ფასი</b>"
            ),
            parse_mode=_PARSE,
        )
        return

    if doc.file_size and doc.file_size > config.MAX_EXCEL_BYTES:
        mb = config.MAX_EXCEL_BYTES // (1024 * 1024)
        await message.bot.send_message(
            chat_id=message.from_user.id,
            text=f"❌ ფაილი ძალიან დიდია. მაქსიმალური ზომა: <b>{mb} MB</b>.",
            parse_mode=_PARSE,
        )
        return

    await message.bot.send_message(
        chat_id=message.from_user.id,
        text="⏳ საწყობის მიღება მუშავდება...",
        parse_mode=_PARSE,
    )

    file_info = await bot.get_file(doc.file_id)
    buf = BytesIO()
    await bot.download_file(file_info.file_path, destination=buf)
    buf.seek(0)

    try:
        wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
        ws = wb.active
    except Exception as exc:
        logger.error("Inventory Excel parse error: %s", exc)
        await message.bot.send_message(
            chat_id=message.from_user.id,
            text=(
                "❌ ფაილი ვერ წაიკითხა. გადაამოწმეთ ფორმატი.\n"
                "სვეტები: <b>თარიღი | OEM | სახელი | მარაგი | ერთეული | ფასი</b>"
            ),
            parse_mode=_PARSE,
        )
        return

    received = 0
    created = 0
    total_value = 0.0
    errors: list[str] = []
    data_rows = 0
    reference = f"xlsx:{doc.file_unique_id}"

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or row[0] is None:
            continue
        data_rows += 1
        if data_rows > _MAX_IMPORT_ROWS:
            errors.append(f"⚠️ ლიმიტი: {_MAX_IMPORT_ROWS} სტრიქონი — დანარჩენი გამოტოვდა.")
            break

        try:
            backdate = _parse_backdate(row[0])
            oem = sanitize_oem(row[1]) if len(row) > 1 else None
            name = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            quantity = float(row[3]) if len(row) > 3 and row[3] is not None else 0.0
            unit = str(row[4]).strip() if len(row) > 4 and row[4] is not None else 'ცალი'
            unit_cost = float(row[5]) if len(row) > 5 and row[5] is not None else 0.0

            if not name:
                continue
            if not oem:
                raise ValueError("OEM კოდი სავალდებულოა — სტრიქონი გამოტოვდა")
            if quantity <= 0:
                raise ValueError("რაოდენობა უნდა იყოს > 0")
            if unit_cost < 0:
                raise ValueError("ფასი უნდა იყოს >= 0")

            result = await db.receive_inventory_batch(
                name=name,
                oem_code=oem,
                quantity=quantity,
                unit_cost=unit_cost,
                min_stock=config.MIN_STOCK_THRESHOLD,
                reference=reference,
                notes=f"Inventory receipt via Excel upload (row {row_idx})",
                received_at=backdate,
                unit=unit,
            )
            received += 1
            total_value += result["total_cost"]
            if result["was_created"]:
                created += 1
        except Exception as exc:
            logger.warning("Inventory row %d error: %s", row_idx, exc)
            error_msg = f"რიგი {row_idx}: {exc}"
            errors.append(error_msg)
            await db.log_parse_failure(
                topic_id=config.STOCK_TOPIC_ID,
                message_text=error_msg,
            )

    summary_lines = [
        "✅ <b>საწყობი განახლდა!</b>",
        f"📊 დამუშავდა: <b>{received}/{data_rows}</b> წარმატებით",
    ]
    if created:
        summary_lines.append(f"🆕 ახალი პროდუქტი: <b>{created}</b>")
    summary_lines.append(f"💰 ჯამური ღირებულება: <b>{total_value:.2f}₾</b>")
    summary_lines.append("📘 ledger: DR 1300 Inventory / CR 2100 Accounts payable")

    if errors:
        summary_lines.append(f"\n❌ <b>ხარვეზები ({len(errors)} რიგი):</b>")
        for err in errors[:5]:
            summary_lines.append(f"  • <code>{html.escape(err)}</code>")
        if len(errors) > 5:
            summary_lines.append(f"  • ... და კიდევ {len(errors) - 5}")
    else:
        summary_lines.append("\n✅ ყველა პროდუქტი წარმატებით აიტვირთულია (ხარვეზების გარეშე)")

    await message.bot.send_message(
        chat_id=message.from_user.id,
        text="\n".join(summary_lines),
        parse_mode=_PARSE,
    )


# ─── Inventory topic: Stock Count / Adjustment upload ─────────────────────────
# Triggered when an admin uploads an Excel file to the Stock topic with a
# caption that starts with "/count" or "#count" (case-insensitive).
#
# Expected columns (header row skipped): OEM | სახელი | ახალი მარაგი
#   OEM         — must match an existing product in the database
#   სახელი      — display name (used in descriptions; DB name is not updated)
#   ახალი მარაგი — the physically-counted quantity (absolute, not a delta)
#
# Logic:
#   delta = target_qty - current_stock
#   delta < 0 → shortage: expense + ledger write-off (DR 7500 / CR 1600)
#   delta > 0 → overage:  stock update + ledger gain  (DR 1600 / CR 7500)
#   delta = 0 → no action
#
# The generated expense has is_non_cash=True so it NEVER reduces the Cash/Bank
# balance — it only moves through the P&L as a write-off.

@sales_router.message(
    InTopic(config.STOCK_TOPIC_ID),
    IsAdmin(),
    F.document,
    F.caption.regexp(r"(?i)^[#/]?count\b"),
)
async def handle_stock_adjustment(message: Message, bot: Bot, db: Database) -> None:
    doc = message.document
    if not doc or not doc.file_name:
        return

    if not doc.file_name.lower().endswith((".xlsx", ".xls")):
        await message.bot.send_message(
            chat_id=message.from_user.id,
            text=(
                "❌ გთხოვთ Excel ფაილი (.xlsx) გამოაგზავნოთ.\n"
                "სვეტები: <b>OEM | სახელი | ახალი მარაგი</b>"
            ),
            parse_mode=_PARSE,
        )
        return

    await message.bot.send_message(
        chat_id=message.from_user.id,
        text="⏳ მარაგის კორექტირება მუშავდება...",
        parse_mode=_PARSE,
    )

    file_info = await bot.get_file(doc.file_id)
    buf = BytesIO()
    await bot.download_file(file_info.file_path, destination=buf)
    buf.seek(0)

    try:
        wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
        ws = wb.active
    except Exception as exc:
        logger.error("Stock count Excel parse error: %s", exc)
        await message.bot.send_message(
            chat_id=message.from_user.id,
            text=(
                "❌ ფაილი ვერ წაიკითხა. გადაამოწმეთ ფორმატი.\n"
                "სვეტები: <b>OEM | სახელი | ახალი მარაგი</b>"
            ),
            parse_mode=_PARSE,
        )
        return

    shortages: list[tuple[str, str, float, float]] = []  # (oem, name, qty, loss_value)
    overages: list[tuple[str, str, float]] = []           # (oem, name, qty)
    unchanged = 0
    errors: list[str] = []
    data_rows = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or row[0] is None:
            continue
        data_rows += 1
        if data_rows > _MAX_IMPORT_ROWS:
            errors.append(f"⚠️ ლიმიტი: {_MAX_IMPORT_ROWS} სტრიქონი — დანარჩენი გამოტოვდა.")
            break

        try:
            oem = sanitize_oem(row[0]) if row[0] is not None else None
            name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            target_qty = float(row[2]) if len(row) > 2 and row[2] is not None else 0.0

            if not oem:
                raise ValueError("OEM კოდი სავალდებულოა")
            if not name:
                raise ValueError("სახელი სავალდებულოა")
            if target_qty < 0:
                raise ValueError("ახალი მარაგი უნდა იყოს >= 0")

            # Fetch current stock and WAC
            product = await db.get_product_by_oem(oem)
            if product is None:
                raise ValueError(f"პროდუქტი OEM '{oem}' ვერ მოიძებნა")

            current_qty = int(product["current_stock"])
            delta = int(target_qty) - current_qty

            if delta == 0:
                unchanged += 1
                continue

            if delta < 0:
                # Shortage: write off the lost units
                wac = await db.get_product_wac(product["id"])
                unit_cost = wac if wac > 0 else float(product["unit_price"])
                result = await db.create_inventory_shortage_expense(
                    oem_code=oem,
                    name=name,
                    shortage_qty=abs(delta),
                    unit_cost=unit_cost,
                )
                shortages.append((oem, name, abs(delta), result["loss_value"]))
            else:
                # Overage: add extra units to stock
                result = await db.record_inventory_overage(
                    oem_code=oem,
                    name=name,
                    overage_qty=delta,
                )
                overages.append((oem, name, delta))

        except Exception as exc:
            logger.warning("Stock count row %d error: %s", row_idx, exc)
            error_msg = f"რიგი {row_idx}: {exc}"
            errors.append(error_msg)
            await db.log_parse_failure(
                topic_id=config.STOCK_TOPIC_ID,
                message_text=error_msg,
            )

    total_loss = sum(v for _, _, _, v in shortages)
    summary_lines = [
        "✅ <b>მარაგის კორექტირება დასრულდა!</b>",
        f"📊 დამუშავდა: <b>{data_rows}</b> სტრიქონი",
    ]
    if unchanged:
        summary_lines.append(f"✔️ უცვლელი: <b>{unchanged}</b>")
    if shortages:
        summary_lines.append(f"\n🔴 <b>დანაკლისი ({len(shortages)} პოზ.):</b>")
        for oem, nm, qty, loss in shortages[:5]:
            summary_lines.append(
                f"  • <code>{html.escape(oem)}</code> {html.escape(nm)} "
                f"— -{qty:.0f} ც. / <b>{loss:.2f}₾</b>"
            )
        if len(shortages) > 5:
            summary_lines.append(f"  • ... და კიდევ {len(shortages) - 5}")
        summary_lines.append(
            f"💸 ჯამური ჩამოწერა: <b>{total_loss:.2f}₾</b> "
            "(P&amp;L ხარჯი — ნაღდ ფულს არ ამცირებს)"
        )
        summary_lines.append("📘 ledger: DR 7500 Write-off / CR 1600 Inventory")
    if overages:
        summary_lines.append(f"\n🟢 <b>ზედმეტობა ({len(overages)} პოზ.):</b>")
        for oem, nm, qty in overages[:5]:
            summary_lines.append(
                f"  • <code>{html.escape(oem)}</code> {html.escape(nm)} — +{qty:.0f} ც."
            )
        if len(overages) > 5:
            summary_lines.append(f"  • ... და კიდევ {len(overages) - 5}")
    if errors:
        summary_lines.append(f"\n❌ <b>ხარვეზები ({len(errors)} რიგი):</b>")
        for err in errors[:5]:
            summary_lines.append(f"  • <code>{html.escape(err)}</code>")
        if len(errors) > 5:
            summary_lines.append(f"  • ... და კიდევ {len(errors) - 5}")

    await message.bot.send_message(
        chat_id=message.from_user.id,
        text="\n".join(summary_lines),
        parse_mode=_PARSE,
    )
